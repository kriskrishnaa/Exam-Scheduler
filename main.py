# -*- coding: utf-8 -*-
import sys
from Tkinter import *
from Tkinter import Tk
from tkFileDialog import askopenfilename
import calendar
from openpyxl import load_workbook,workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.pagesizes import landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab import rl_config
from reportlab.lib import utils
global canvas
#hall_list = ['IE1101','IE1102','IE1103','IE1104','IE1105','IE1106','IE1107','IS2101','IS2102','IS2103','IS2104','IS2105','IW2101','IW2102','IW2103','IW2104','IW2105','IW2106','IW2107','IW2108','IW2109','IW2110','IW2111','IW2112']
with open('college_hall_list.txt', 'r') as myfile:
    clge_hall=myfile.read().replace('\n', '')
clge_hall = str(clge_hall)
hall_list = clge_hall.split(" ")
path2 = ""
stack = []
def current_ha(hall_no):
    global current_hall
    current_hall = hall_no
def hall_n():
    global floor
    floor = ""
    if hall_no[2] == "1":
        floor = "(Ground Floor)"
    elif hall_no[2] == "2":
        floor = "(First Floor)"
    elif hall_no[2] == "3":
        floor = "(Second Floor)"
    elif hall_no[2] == "4":
        floor = "(Third Floor)"
def atlast(dept1,dept2):
    #print "came in atlast"
    global stack
    stack.append(dept1)
    stack.append(dept2)
    #print "##########stack-%s" % (stack)
    #print len(stack)
    #canvas.save()
def atlastmerge(canvas):
    with open('college_name.txt', 'r') as myfile:
        clge_name=myfile.read().replace('\n', '')
    clge_name = str(clge_name)
    from compiler.ast import flatten

    
    hall_no = current_hall
    hall_index = 0
    hall_index = hall_list.index(hall_no)
    hall_n()
    hallno = "Hall No - %s %s" % (hall_no,floor)
    max_ele = len(flatten(stack))
    f=0
    stackn = []
    #for f in range(len(stack)):
        #stackn.append(stack[f])
    
    stackn = reduce(lambda x,y: x+y,stack)
    ##print stack
    #print "###################atlast"
    #print max_ele
    #print len(stackn)
    l3 = max_ele / 25
    #print "stackn--------%s" % (stackn)
    if l3 == 0:
        p = 0
        for p in range(25-(max_ele)):
            stackn.append('')
        pdfmetrics.registerFont(TTFont('Bookman Old Style', 'BOOKOSB.ttf'))
        pdfmetrics.registerFont(TTFont('Palatino Linotype', 'palatino-linotype-bold-italic.ttf'))
        pdfmetrics.registerFont(TTFont('Times New Roman', 'Times_Normal.ttf'))
        canvas.setFont('Times New Roman', 18)
        canvas.drawString(200.83,575.8,clge_name)
        canvas.drawString(200.83,550.8,"Anna University Examinations - ")
        canvas.drawString(457.83,550.8,month)
            
            
        canvas.setFont('Bookman Old Style', 23)
        canvas.drawString(369.67,458.8,hallno)
        canvas.setFont('Palatino Linotype', 18)
        canvas.drawString(456.33,490.40,"Session: ")
        canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                
        canvas.drawString(90.83,490.40,'Date: ')
        canvas.drawString(150.83,490.40,date)
        canvas.drawString(90.83,458.8,'No. of Candidates: ')
        canvas.drawString(250.83,458.8,str(max_ele))
        canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
        canvas.setFont('Palatino Linotype', 15)
            
        canvas.line(31.83,133,765.92,133) #"""bottom most line"""
        canvas.line(31.83,133,31.83,419) #"""left most line"""
        canvas.line(31.83,419,765.92,419) #"""top most line"""
        canvas.line(765.92,133,765.92,419) #"""right most line"""
        canvas.line(80.23,133,80.23,419) #"""1st column line"""
        canvas.line(179.59,133,179.59,419) #"""2nd column line"""
        canvas.line(226.39,133,226.39,419) #"""3rd column line"""
        canvas.line(324.29,133,324.29,419) #""" 4th column line"""
        canvas.line(371.09,133,371.09,419) #""" 5th column line"""
        canvas.line(473.29,133,473.29,419) #""" 6th column line"""
        canvas.line(520.09,133,520.09,419) #""" 7th column line"""
        canvas.line(621.67,133,621.67,419) #""" 8th column line"""
        canvas.line(667.03,133,667.03,419) #""" 9th column line"""
            
        canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
        canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
        canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
        canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
        canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
            
        canvas.drawString(36.83,401.67,'Table')
        canvas.drawString(42.83,386.67,'No.')
        canvas.drawString(103.64,394.17,'Reg.No')
        canvas.drawString(184.59,401.67,'Table')
        canvas.drawString(190.59,386.67,'No.')
        canvas.drawString(249.8,394.17,'Reg.No')
        canvas.drawString(329.29,401.67,'Table')
        canvas.drawString(335.29,386.67,'No.')
        canvas.drawString(394.5,394.17,'Reg.No')
        canvas.drawString(478.29,401.67,'Table')
        canvas.drawString(484.29,386.67,'No.')
        canvas.drawString(543.5,394.67,'Reg.No')
        canvas.drawString(626.67,401.67,'Table')
        canvas.drawString(632.67,386.67,'No.')
        canvas.drawString(690.44,394.67,'Reg.No')
        r1 = [0,1,2,5]
        canvas.drawString(51.83,349.936,'1')
        canvas.drawString(51.83,300.616,'2')
        canvas.drawString(51.83,251.296,'3')
        canvas.drawString(51.83,201.686,'4')
        canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
            
        canvas.drawString(194.59,349.936,'10')
        canvas.drawString(199.59,300.616,'9')
        canvas.drawString(199.59,251.296,'8')
        canvas.drawString(199.59,201.686,'7')
        canvas.drawString(199.59,152.076,'6')
            
        canvas.drawString(339.29,349.936,'11')
        canvas.drawString(339.29,300.616,'12')
        canvas.drawString(339.29,251.296,'13')
        canvas.drawString(339.29,201.686,'14')
        canvas.drawString(339.29,152.076,'15')
            
        canvas.drawString(488.29,349.936,'20')
        canvas.drawString(488.29,300.616,'19')
        canvas.drawString(488.29,251.296,'18')
        canvas.drawString(488.29,201.686,'17')
        canvas.drawString(488.29,152.076,'16')
            
        canvas.drawString(636.67,349.936,'21')
        canvas.drawString(636.67,300.616,'22')
        canvas.drawString(636.67,251.296,'23')
        canvas.drawString(636.67,201.686,'24')
        canvas.drawString(636.67,152.076,'25')
        canvas.drawString(84.846,349.936,str(stackn[0]).decode('utf-8'))
        canvas.drawString(84.846,300.616,str(stackn[1]).decode('utf-8'))
        canvas.drawString(84.846,251.296,str(stackn[2]).decode('utf-8'))
        canvas.drawString(84.846,201.686,str(stackn[3]).decode('utf-8'))
        canvas.drawString(84.846,152.076,str(stackn[4]).decode('utf-8'))
                
        canvas.drawString(231.006,349.936,str(stackn[9]).decode('utf-8'))
        canvas.drawString(231.006,300.616,str(stackn[8]).decode('utf-8'))
        canvas.drawString(231.006,251.296,str(stackn[7]).decode('utf-8'))
        canvas.drawString(231.006,201.686,str(stackn[6]).decode('utf-8'))
        canvas.drawString(231.006,152.076,str(stackn[5]).decode('utf-8'))
                
        canvas.drawString(375.706,349.936,str(stackn[10]).decode('utf-8'))
        canvas.drawString(375.706,300.616,str(stackn[11]).decode('utf-8'))
        canvas.drawString(375.706,251.296,str(stackn[12]).decode('utf-8'))
        canvas.drawString(375.706,201.686,str(stackn[13]).decode('utf-8'))
        canvas.drawString(375.706,152.076,str(stackn[14]).decode('utf-8'))
                
        canvas.drawString(524.706,349.936,str(stackn[19]).decode('utf-8'))
        canvas.drawString(524.706,300.616,str(stackn[18]).decode('utf-8'))
        canvas.drawString(524.706,251.296,str(stackn[17]).decode('utf-8'))
        canvas.drawString(524.706,201.686,str(stackn[16]).decode('utf-8'))
        canvas.drawString(524.706,152.076,str(stackn[15]).decode('utf-8'))
                
        canvas.drawString(671.646,349.936,str(stackn[20]).decode('utf-8'))
        canvas.drawString(671.646,300.616,str(stackn[21]).decode('utf-8'))
        canvas.drawString(671.646,251.296,str(stackn[22]).decode('utf-8'))
        canvas.drawString(671.646,201.686,str(stackn[23]).decode('utf-8'))
        canvas.drawString(671.646,152.076,str(stackn[24]).decode('utf-8'))
        canvas.showPage()
    elif l3 == 1:
        p = 0
        for p in range(50-(max_ele)):
            stackn.append('')
        pdfmetrics.registerFont(TTFont('Bookman Old Style', 'BOOKOSB.ttf'))
        pdfmetrics.registerFont(TTFont('Palatino Linotype', 'palatino-linotype-bold-italic.ttf'))
        pdfmetrics.registerFont(TTFont('Times New Roman', 'Times_Normal.ttf'))
        canvas.setFont('Times New Roman', 18)
        canvas.drawString(200.83,575.8,clge_name)
        canvas.drawString(200.83,550.8,"Anna University Examinations - ")
        canvas.drawString(457.83,550.8,month)
            
            
        canvas.setFont('Bookman Old Style', 23)
        canvas.drawString(369.67,458.8,hallno)
        canvas.setFont('Palatino Linotype', 18)
        canvas.drawString(456.33,490.40,"Session: ")
        canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                
        canvas.drawString(90.83,490.40,'Date: ')
        canvas.drawString(150.83,490.40,date)
        canvas.drawString(90.83,458.8,'No. of Candidates: ')
        canvas.drawString(250.83,458.8,"25")
        canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
        canvas.setFont('Palatino Linotype', 15)
            
        canvas.line(31.83,133,765.92,133) #"""bottom most line"""
        canvas.line(31.83,133,31.83,419) #"""left most line"""
        canvas.line(31.83,419,765.92,419) #"""top most line"""
        canvas.line(765.92,133,765.92,419) #"""right most line"""
        canvas.line(80.23,133,80.23,419) #"""1st column line"""
        canvas.line(179.59,133,179.59,419) #"""2nd column line"""
        canvas.line(226.39,133,226.39,419) #"""3rd column line"""
        canvas.line(324.29,133,324.29,419) #""" 4th column line"""
        canvas.line(371.09,133,371.09,419) #""" 5th column line"""
        canvas.line(473.29,133,473.29,419) #""" 6th column line"""
        canvas.line(520.09,133,520.09,419) #""" 7th column line"""
        canvas.line(621.67,133,621.67,419) #""" 8th column line"""
        canvas.line(667.03,133,667.03,419) #""" 9th column line"""
            
        canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
        canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
        canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
        canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
        canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
            
        canvas.drawString(36.83,401.67,'Table')
        canvas.drawString(42.83,386.67,'No.')
        canvas.drawString(103.64,394.17,'Reg.No')
        canvas.drawString(184.59,401.67,'Table')
        canvas.drawString(190.59,386.67,'No.')
        canvas.drawString(249.8,394.17,'Reg.No')
        canvas.drawString(329.29,401.67,'Table')
        canvas.drawString(335.29,386.67,'No.')
        canvas.drawString(394.5,394.17,'Reg.No')
        canvas.drawString(478.29,401.67,'Table')
        canvas.drawString(484.29,386.67,'No.')
        canvas.drawString(543.5,394.67,'Reg.No')
        canvas.drawString(626.67,401.67,'Table')
        canvas.drawString(632.67,386.67,'No.')
        canvas.drawString(690.44,394.67,'Reg.No')
        r1 = [0,1,2,5]
        canvas.drawString(51.83,349.936,'1')
        canvas.drawString(51.83,300.616,'2')
        canvas.drawString(51.83,251.296,'3')
        canvas.drawString(51.83,201.686,'4')
        canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
            
        canvas.drawString(194.59,349.936,'10')
        canvas.drawString(199.59,300.616,'9')
        canvas.drawString(199.59,251.296,'8')
        canvas.drawString(199.59,201.686,'7')
        canvas.drawString(199.59,152.076,'6')
            
        canvas.drawString(339.29,349.936,'11')
        canvas.drawString(339.29,300.616,'12')
        canvas.drawString(339.29,251.296,'13')
        canvas.drawString(339.29,201.686,'14')
        canvas.drawString(339.29,152.076,'15')
            
        canvas.drawString(488.29,349.936,'20')
        canvas.drawString(488.29,300.616,'19')
        canvas.drawString(488.29,251.296,'18')
        canvas.drawString(488.29,201.686,'17')
        canvas.drawString(488.29,152.076,'16')
            
        canvas.drawString(636.67,349.936,'21')
        canvas.drawString(636.67,300.616,'22')
        canvas.drawString(636.67,251.296,'23')
        canvas.drawString(636.67,201.686,'24')
        canvas.drawString(636.67,152.076,'25')
        canvas.drawString(84.846,349.936,str(stackn[0]).decode('utf-8'))
        canvas.drawString(84.846,300.616,str(stackn[1]).decode('utf-8'))
        canvas.drawString(84.846,251.296,str(stackn[2]).decode('utf-8'))
        canvas.drawString(84.846,201.686,str(stackn[3]).decode('utf-8'))
        canvas.drawString(84.846,152.076,str(stackn[4]).decode('utf-8'))
                
        canvas.drawString(231.006,349.936,str(stackn[9]).decode('utf-8'))
        canvas.drawString(231.006,300.616,str(stackn[8]).decode('utf-8'))
        canvas.drawString(231.006,251.296,str(stackn[7]).decode('utf-8'))
        canvas.drawString(231.006,201.686,str(stackn[6]).decode('utf-8'))
        canvas.drawString(231.006,152.076,str(stackn[5]).decode('utf-8'))
                
        canvas.drawString(375.706,349.936,str(stackn[10]).decode('utf-8'))
        canvas.drawString(375.706,300.616,str(stackn[11]).decode('utf-8'))
        canvas.drawString(375.706,251.296,str(stackn[12]).decode('utf-8'))
        canvas.drawString(375.706,201.686,str(stackn[13]).decode('utf-8'))
        canvas.drawString(375.706,152.076,str(stackn[14]).decode('utf-8'))
                
        canvas.drawString(524.706,349.936,str(stackn[19]).decode('utf-8'))
        canvas.drawString(524.706,300.616,str(stackn[18]).decode('utf-8'))
        canvas.drawString(524.706,251.296,str(stackn[17]).decode('utf-8'))
        canvas.drawString(524.706,201.686,str(stackn[16]).decode('utf-8'))
        canvas.drawString(524.706,152.076,str(stackn[15]).decode('utf-8'))
                
        canvas.drawString(671.646,349.936,str(stackn[20]).decode('utf-8'))
        canvas.drawString(671.646,300.616,str(stackn[21]).decode('utf-8'))
        canvas.drawString(671.646,251.296,str(stackn[22]).decode('utf-8'))
        canvas.drawString(671.646,201.686,str(stackn[23]).decode('utf-8'))
        canvas.drawString(671.646,152.076,str(stackn[24]).decode('utf-8'))
        canvas.showPage()
        hall_index += 1
        hall_no = hall_list[hall_index]
        hall_n()
        hallno = "Hall No - %s %s" % (hall_no,floor)
        pdfmetrics.registerFont(TTFont('Bookman Old Style', 'BOOKOSB.ttf'))
        pdfmetrics.registerFont(TTFont('Palatino Linotype', 'palatino-linotype-bold-italic.ttf'))
        pdfmetrics.registerFont(TTFont('Times New Roman', 'Times_Normal.ttf'))
        canvas.setFont('Times New Roman', 18)
        canvas.drawString(200.83,575.8,clge_name)
        canvas.drawString(200.83,550.8,"Anna University Examinations - ")
        canvas.drawString(457.83,550.8,month)
            
            
        canvas.setFont('Bookman Old Style', 23)
        canvas.drawString(369.67,458.8,hallno)
        canvas.setFont('Palatino Linotype', 18)
        canvas.drawString(456.33,490.40,"Session: ")
        canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                
        canvas.drawString(90.83,490.40,'Date: ')
        canvas.drawString(150.83,490.40,date)
        canvas.drawString(90.83,458.8,'No. of Candidates: ')
        no_ofstudents = 0
        no_ofstudents = max_ele - 25
        canvas.drawString(250.83,458.8,str(no_ofstudents))
        canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
        canvas.setFont('Palatino Linotype', 15)
            
        canvas.line(31.83,133,765.92,133) #"""bottom most line"""
        canvas.line(31.83,133,31.83,419) #"""left most line"""
        canvas.line(31.83,419,765.92,419) #"""top most line"""
        canvas.line(765.92,133,765.92,419) #"""right most line"""
        canvas.line(80.23,133,80.23,419) #"""1st column line"""
        canvas.line(179.59,133,179.59,419) #"""2nd column line"""
        canvas.line(226.39,133,226.39,419) #"""3rd column line"""
        canvas.line(324.29,133,324.29,419) #""" 4th column line"""
        canvas.line(371.09,133,371.09,419) #""" 5th column line"""
        canvas.line(473.29,133,473.29,419) #""" 6th column line"""
        canvas.line(520.09,133,520.09,419) #""" 7th column line"""
        canvas.line(621.67,133,621.67,419) #""" 8th column line"""
        canvas.line(667.03,133,667.03,419) #""" 9th column line"""
            
        canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
        canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
        canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
        canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
        canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
            
        canvas.drawString(36.83,401.67,'Table')
        canvas.drawString(42.83,386.67,'No.')
        canvas.drawString(103.64,394.17,'Reg.No')
        canvas.drawString(184.59,401.67,'Table')
        canvas.drawString(190.59,386.67,'No.')
        canvas.drawString(249.8,394.17,'Reg.No')
        canvas.drawString(329.29,401.67,'Table')
        canvas.drawString(335.29,386.67,'No.')
        canvas.drawString(394.5,394.17,'Reg.No')
        canvas.drawString(478.29,401.67,'Table')
        canvas.drawString(484.29,386.67,'No.')
        canvas.drawString(543.5,394.67,'Reg.No')
        canvas.drawString(626.67,401.67,'Table')
        canvas.drawString(632.67,386.67,'No.')
        canvas.drawString(690.44,394.67,'Reg.No')
        r1 = [0,1,2,5]
        canvas.drawString(51.83,349.936,'1')
        canvas.drawString(51.83,300.616,'2')
        canvas.drawString(51.83,251.296,'3')
        canvas.drawString(51.83,201.686,'4')
        canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
            
        canvas.drawString(194.59,349.936,'10')
        canvas.drawString(199.59,300.616,'9')
        canvas.drawString(199.59,251.296,'8')
        canvas.drawString(199.59,201.686,'7')
        canvas.drawString(199.59,152.076,'6')
            
        canvas.drawString(339.29,349.936,'11')
        canvas.drawString(339.29,300.616,'12')
        canvas.drawString(339.29,251.296,'13')
        canvas.drawString(339.29,201.686,'14')
        canvas.drawString(339.29,152.076,'15')
            
        canvas.drawString(488.29,349.936,'20')
        canvas.drawString(488.29,300.616,'19')
        canvas.drawString(488.29,251.296,'18')
        canvas.drawString(488.29,201.686,'17')
        canvas.drawString(488.29,152.076,'16')
            
        canvas.drawString(636.67,349.936,'21')
        canvas.drawString(636.67,300.616,'22')
        canvas.drawString(636.67,251.296,'23')
        canvas.drawString(636.67,201.686,'24')
        canvas.drawString(636.67,152.076,'25')
        canvas.drawString(84.846,349.936,str(stackn[25]).decode('utf-8'))
        canvas.drawString(84.846,300.616,str(stackn[26]).decode('utf-8'))
        canvas.drawString(84.846,251.296,str(stackn[27]).decode('utf-8'))
        canvas.drawString(84.846,201.686,str(stackn[28]).decode('utf-8'))
        canvas.drawString(84.846,152.076,str(stackn[29]).decode('utf-8'))
                
        canvas.drawString(231.006,349.936,str(stackn[34]).decode('utf-8'))
        canvas.drawString(231.006,300.616,str(stackn[33]).decode('utf-8'))
        canvas.drawString(231.006,251.296,str(stackn[32]).decode('utf-8'))
        canvas.drawString(231.006,201.686,str(stackn[31]).decode('utf-8'))
        canvas.drawString(231.006,152.076,str(stackn[30]).decode('utf-8'))
                
        canvas.drawString(375.706,349.936,str(stackn[35]).decode('utf-8'))
        canvas.drawString(375.706,300.616,str(stackn[36]).decode('utf-8'))
        canvas.drawString(375.706,251.296,str(stackn[37]).decode('utf-8'))
        canvas.drawString(375.706,201.686,str(stackn[38]).decode('utf-8'))
        canvas.drawString(375.706,152.076,str(stackn[39]).decode('utf-8'))
                
        canvas.drawString(524.706,349.936,str(stackn[44]).decode('utf-8'))
        canvas.drawString(524.706,300.616,str(stackn[43]).decode('utf-8'))
        canvas.drawString(524.706,251.296,str(stackn[42]).decode('utf-8'))
        canvas.drawString(524.706,201.686,str(stackn[41]).decode('utf-8'))
        canvas.drawString(524.706,152.076,str(stackn[40]).decode('utf-8'))
                
        canvas.drawString(671.646,349.936,str(stackn[45]).decode('utf-8'))
        canvas.drawString(671.646,300.616,str(stackn[46]).decode('utf-8'))
        canvas.drawString(671.646,251.296,str(stackn[47]).decode('utf-8'))
        canvas.drawString(671.646,201.686,str(stackn[48]).decode('utf-8'))
        canvas.drawString(671.646,152.076,str(stackn[49]).decode('utf-8'))
        canvas.showPage()

    else:
        s = 0
        while s != l3:
            pdfmetrics.registerFont(TTFont('Bookman Old Style', 'BOOKOSB.ttf'))
            pdfmetrics.registerFont(TTFont('Palatino Linotype', 'palatino-linotype-bold-italic.ttf'))
            pdfmetrics.registerFont(TTFont('Times New Roman', 'Times_Normal.ttf'))
            canvas.setFont('Times New Roman', 18)
            canvas.drawString(200.83,575.8,clge_name)
            canvas.drawString(200.83,550.8,"Anna University Examinations - ")
            canvas.drawString(457.83,550.8,month)
                
                
            canvas.setFont('Bookman Old Style', 23)
            canvas.drawString(369.67,458.8,hallno)
            canvas.setFont('Palatino Linotype', 18)
            canvas.drawString(456.33,490.40,"Session: ")
            canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                    
            canvas.drawString(90.83,490.40,'Date: ')
            canvas.drawString(150.83,490.40,date)
            canvas.drawString(90.83,458.8,'No. of Candidates: ')
            canvas.drawString(250.83,458.8,"25")
            canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
            canvas.setFont('Palatino Linotype', 15)
                
            canvas.line(31.83,133,765.92,133) #"""bottom most line"""
            canvas.line(31.83,133,31.83,419) #"""left most line"""
            canvas.line(31.83,419,765.92,419) #"""top most line"""
            canvas.line(765.92,133,765.92,419) #"""right most line"""
            canvas.line(80.23,133,80.23,419) #"""1st column line"""
            canvas.line(179.59,133,179.59,419) #"""2nd column line"""
            canvas.line(226.39,133,226.39,419) #"""3rd column line"""
            canvas.line(324.29,133,324.29,419) #""" 4th column line"""
            canvas.line(371.09,133,371.09,419) #""" 5th column line"""
            canvas.line(473.29,133,473.29,419) #""" 6th column line"""
            canvas.line(520.09,133,520.09,419) #""" 7th column line"""
            canvas.line(621.67,133,621.67,419) #""" 8th column line"""
            canvas.line(667.03,133,667.03,419) #""" 9th column line"""
                
            canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
            canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
            canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
            canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
            canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
                
            canvas.drawString(36.83,401.67,'Table')
            canvas.drawString(42.83,386.67,'No.')
            canvas.drawString(103.64,394.17,'Reg.No')
            canvas.drawString(184.59,401.67,'Table')
            canvas.drawString(190.59,386.67,'No.')
            canvas.drawString(249.8,394.17,'Reg.No')
            canvas.drawString(329.29,401.67,'Table')
            canvas.drawString(335.29,386.67,'No.')
            canvas.drawString(394.5,394.17,'Reg.No')
            canvas.drawString(478.29,401.67,'Table')
            canvas.drawString(484.29,386.67,'No.')
            canvas.drawString(543.5,394.67,'Reg.No')
            canvas.drawString(626.67,401.67,'Table')
            canvas.drawString(632.67,386.67,'No.')
            canvas.drawString(690.44,394.67,'Reg.No')
            r1 = [0,1,2,5]
            canvas.drawString(51.83,349.936,'1')
            canvas.drawString(51.83,300.616,'2')
            canvas.drawString(51.83,251.296,'3')
            canvas.drawString(51.83,201.686,'4')
            canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
                
            canvas.drawString(194.59,349.936,'10')
            canvas.drawString(199.59,300.616,'9')
            canvas.drawString(199.59,251.296,'8')
            canvas.drawString(199.59,201.686,'7')
            canvas.drawString(199.59,152.076,'6')
                
            canvas.drawString(339.29,349.936,'11')
            canvas.drawString(339.29,300.616,'12')
            canvas.drawString(339.29,251.296,'13')
            canvas.drawString(339.29,201.686,'14')
            canvas.drawString(339.29,152.076,'15')
                
            canvas.drawString(488.29,349.936,'20')
            canvas.drawString(488.29,300.616,'19')
            canvas.drawString(488.29,251.296,'18')
            canvas.drawString(488.29,201.686,'17')
            canvas.drawString(488.29,152.076,'16')
                
            canvas.drawString(636.67,349.936,'21')
            canvas.drawString(636.67,300.616,'22')
            canvas.drawString(636.67,251.296,'23')
            canvas.drawString(636.67,201.686,'24')
            canvas.drawString(636.67,152.076,'25')
            canvas.drawString(84.846,349.936,str(stackn[0]).decode('utf-8'))
            canvas.drawString(84.846,300.616,str(stackn[1]).decode('utf-8'))
            canvas.drawString(84.846,251.296,str(stackn[2]).decode('utf-8'))
            canvas.drawString(84.846,201.686,str(stackn[3]).decode('utf-8'))
            canvas.drawString(84.846,152.076,str(stackn[4]).decode('utf-8'))
                    
            canvas.drawString(231.006,349.936,str(stackn[9]).decode('utf-8'))
            canvas.drawString(231.006,300.616,str(stackn[8]).decode('utf-8'))
            canvas.drawString(231.006,251.296,str(stackn[7]).decode('utf-8'))
            canvas.drawString(231.006,201.686,str(stackn[6]).decode('utf-8'))
            canvas.drawString(231.006,152.076,str(stackn[5]).decode('utf-8'))
                    
            canvas.drawString(375.706,349.936,str(stackn[10]).decode('utf-8'))
            canvas.drawString(375.706,300.616,str(stackn[11]).decode('utf-8'))
            canvas.drawString(375.706,251.296,str(stackn[12]).decode('utf-8'))
            canvas.drawString(375.706,201.686,str(stackn[13]).decode('utf-8'))
            canvas.drawString(375.706,152.076,str(stackn[14]).decode('utf-8'))
                    
            canvas.drawString(524.706,349.936,str(stackn[19]).decode('utf-8'))
            canvas.drawString(524.706,300.616,str(stackn[18]).decode('utf-8'))
            canvas.drawString(524.706,251.296,str(stackn[17]).decode('utf-8'))
            canvas.drawString(524.706,201.686,str(stackn[16]).decode('utf-8'))
            canvas.drawString(524.706,152.076,str(stackn[15]).decode('utf-8'))
                    
            canvas.drawString(671.646,349.936,str(stackn[20]).decode('utf-8'))
            canvas.drawString(671.646,300.616,str(stackn[21]).decode('utf-8'))
            canvas.drawString(671.646,251.296,str(stackn[22]).decode('utf-8'))
            canvas.drawString(671.646,201.686,str(stackn[23]).decode('utf-8'))
            canvas.drawString(671.646,152.076,str(stackn[24]).decode('utf-8'))
            canvas.showPage()
            hall_index += 1
            hall_no = hall_list[hall_index]
            hall_n()
            hallno = "Hall No - %s %s" % (hall_no,floor)
            s = s + 1
            try:
                
                for b in range(25):
                    stackn.pop(0)
            except IndexError:
                pass
            if s == l3:
                for k in range(25-len(stackn)):
                    stackn.append("")
                pdfmetrics.registerFont(TTFont('Bookman Old Style', 'BOOKOSB.ttf'))
                pdfmetrics.registerFont(TTFont('Palatino Linotype', 'palatino-linotype-bold-italic.ttf'))
                pdfmetrics.registerFont(TTFont('Times New Roman', 'Times_Normal.ttf'))
                canvas.setFont('Times New Roman', 18)
                canvas.drawString(200.83,575.8,clge_name)
                canvas.drawString(200.83,550.8,"Anna University Examinations - ")
                canvas.drawString(457.83,550.8,month)
                    
                    
                canvas.setFont('Bookman Old Style', 23)
                canvas.drawString(369.67,458.8,hallno)
                canvas.setFont('Palatino Linotype', 18)
                canvas.drawString(456.33,490.40,"Session: ")
                canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                        
                canvas.drawString(90.83,490.40,'Date: ')
                canvas.drawString(150.83,490.40,date)
                canvas.drawString(90.83,458.8,'No. of Candidates: ')
                no_of_students = 0
                no_of_students = max_ele - (l3 * 25)
                canvas.drawString(250.83,458.8,str(no_of_students))
                canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
                canvas.setFont('Palatino Linotype', 15)
                    
                canvas.line(31.83,133,765.92,133) #"""bottom most line"""
                canvas.line(31.83,133,31.83,419) #"""left most line"""
                canvas.line(31.83,419,765.92,419) #"""top most line"""
                canvas.line(765.92,133,765.92,419) #"""right most line"""
                canvas.line(80.23,133,80.23,419) #"""1st column line"""
                canvas.line(179.59,133,179.59,419) #"""2nd column line"""
                canvas.line(226.39,133,226.39,419) #"""3rd column line"""
                canvas.line(324.29,133,324.29,419) #""" 4th column line"""
                canvas.line(371.09,133,371.09,419) #""" 5th column line"""
                canvas.line(473.29,133,473.29,419) #""" 6th column line"""
                canvas.line(520.09,133,520.09,419) #""" 7th column line"""
                canvas.line(621.67,133,621.67,419) #""" 8th column line"""
                canvas.line(667.03,133,667.03,419) #""" 9th column line"""
                    
                canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
                canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
                canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
                canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
                canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
                    
                canvas.drawString(36.83,401.67,'Table')
                canvas.drawString(42.83,386.67,'No.')
                canvas.drawString(103.64,394.17,'Reg.No')
                canvas.drawString(184.59,401.67,'Table')
                canvas.drawString(190.59,386.67,'No.')
                canvas.drawString(249.8,394.17,'Reg.No')
                canvas.drawString(329.29,401.67,'Table')
                canvas.drawString(335.29,386.67,'No.')
                canvas.drawString(394.5,394.17,'Reg.No')
                canvas.drawString(478.29,401.67,'Table')
                canvas.drawString(484.29,386.67,'No.')
                canvas.drawString(543.5,394.67,'Reg.No')
                canvas.drawString(626.67,401.67,'Table')
                canvas.drawString(632.67,386.67,'No.')
                canvas.drawString(690.44,394.67,'Reg.No')
                r1 = [0,1,2,5]
                canvas.drawString(51.83,349.936,'1')
                canvas.drawString(51.83,300.616,'2')
                canvas.drawString(51.83,251.296,'3')
                canvas.drawString(51.83,201.686,'4')
                canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
                    
                canvas.drawString(194.59,349.936,'10')
                canvas.drawString(199.59,300.616,'9')
                canvas.drawString(199.59,251.296,'8')
                canvas.drawString(199.59,201.686,'7')
                canvas.drawString(199.59,152.076,'6')
                    
                canvas.drawString(339.29,349.936,'11')
                canvas.drawString(339.29,300.616,'12')
                canvas.drawString(339.29,251.296,'13')
                canvas.drawString(339.29,201.686,'14')
                canvas.drawString(339.29,152.076,'15')
                    
                canvas.drawString(488.29,349.936,'20')
                canvas.drawString(488.29,300.616,'19')
                canvas.drawString(488.29,251.296,'18')
                canvas.drawString(488.29,201.686,'17')
                canvas.drawString(488.29,152.076,'16')
                    
                canvas.drawString(636.67,349.936,'21')
                canvas.drawString(636.67,300.616,'22')
                canvas.drawString(636.67,251.296,'23')
                canvas.drawString(636.67,201.686,'24')
                canvas.drawString(636.67,152.076,'25')
                canvas.drawString(84.846,349.936,str(stackn[0]).decode('utf-8'))
                canvas.drawString(84.846,300.616,str(stackn[1]).decode('utf-8'))
                canvas.drawString(84.846,251.296,str(stackn[2]).decode('utf-8'))
                canvas.drawString(84.846,201.686,str(stackn[3]).decode('utf-8'))
                canvas.drawString(84.846,152.076,str(stackn[4]).decode('utf-8'))
                        
                canvas.drawString(231.006,349.936,str(stackn[9]).decode('utf-8'))
                canvas.drawString(231.006,300.616,str(stackn[8]).decode('utf-8'))
                canvas.drawString(231.006,251.296,str(stackn[7]).decode('utf-8'))
                canvas.drawString(231.006,201.686,str(stackn[6]).decode('utf-8'))
                canvas.drawString(231.006,152.076,str(stackn[5]).decode('utf-8'))
                        
                canvas.drawString(375.706,349.936,str(stackn[10]).decode('utf-8'))
                canvas.drawString(375.706,300.616,str(stackn[11]).decode('utf-8'))
                canvas.drawString(375.706,251.296,str(stackn[12]).decode('utf-8'))
                canvas.drawString(375.706,201.686,str(stackn[13]).decode('utf-8'))
                canvas.drawString(375.706,152.076,str(stackn[14]).decode('utf-8'))
                        
                canvas.drawString(524.706,349.936,str(stackn[19]).decode('utf-8'))
                canvas.drawString(524.706,300.616,str(stackn[18]).decode('utf-8'))
                canvas.drawString(524.706,251.296,str(stackn[17]).decode('utf-8'))
                canvas.drawString(524.706,201.686,str(stackn[16]).decode('utf-8'))
                canvas.drawString(524.706,152.076,str(stackn[15]).decode('utf-8'))
                        
                canvas.drawString(671.646,349.936,str(stackn[20]).decode('utf-8'))
                canvas.drawString(671.646,300.616,str(stackn[21]).decode('utf-8'))
                canvas.drawString(671.646,251.296,str(stackn[22]).decode('utf-8'))
                canvas.drawString(671.646,201.686,str(stackn[23]).decode('utf-8'))
                canvas.drawString(671.646,152.076,str(stackn[24]).decode('utf-8'))
                canvas.showPage()
                

        
    
    

def single_dept(hall_no,canvas):
    with open('college_name.txt', 'r') as myfile:
        clge_name=myfile.read().replace('\n', '')
    clge_name = str(clge_name)
    #print "sinledept called"
    #dept1 = [' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',' ',] 
    dept1 = []
    book1 = load_workbook("output1.xlsx")
    dept_01 = book1.get_active_sheet()
    rowct1 = 6
    i1 = 6
    loop1 = 0
    sheet1 = book1.worksheets[0]
    row_count1 = sheet1.max_row
    hall_index = 0
    hall_index = hall_list.index(hall_no)
    hall_n()
    hallno = "Hall No - %s %s" % (hall_no,floor)
    while rowct1 < row_count1:
        try:
            while str(dept_01.cell(row = i1, column = 1).value[0:7]) != "College":
                i1 += 1
                loop1 += 1
                
        except TypeError:
            pass
            
        #print loop1
        r1 = 0
        for r1 in range(loop1):
            if len(str(dept_01.cell(row = rowct1, column = 1).value)) == 14:
                dept1.append(str(dept_01.cell(row = rowct1, column = 1).value[2:14]))
            elif len(str(dept_01.cell(row = rowct1, column = 1).value)) == 15:
                dept1.append(str(dept_01.cell(row = rowct1, column = 1).value[3:15]))
            else:
                dept1.append(str(dept_01.cell(row = rowct1, column = 1).value[4:16]))
            rowct1 = rowct1 + 1
        
        rowct1 = rowct1 + 5
        i1 = rowct1
        loop1 = 0
    max_ele = len(dept1)
    l = len(dept1)/25
    i = 0
    #print dept1
    #month = "NOV/DEC"
    #date = "10-12-17"
    #hallno = "IN2512"
    #session = "FN"
    while i != l:
        
        pdfmetrics.registerFont(TTFont('Bookman Old Style', 'BOOKOSB.ttf'))
        pdfmetrics.registerFont(TTFont('Palatino Linotype', 'palatino-linotype-bold-italic.ttf'))
        pdfmetrics.registerFont(TTFont('Times New Roman', 'Times_Normal.ttf'))
        canvas.setFont('Times New Roman', 18)
        canvas.drawString(200.83,575.8,clge_name)
        canvas.drawString(200.83,550.8,"Anna University Examinations - ")
        canvas.drawString(457.83,550.8,month)
             
        canvas.setFont('Bookman Old Style', 23)
        canvas.drawString(369.67,458.8,hallno)
        canvas.setFont('Palatino Linotype', 18)
        canvas.drawString(456.33,490.40,"Session: ")
        canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                
        canvas.drawString(90.83,490.40,'Date: ')
        canvas.drawString(150.83,490.40,date)
        canvas.drawString(90.83,458.8,'No. of Candidates: ')
        canvas.drawString(250.83,458.8,"25")
        canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
        canvas.setFont('Palatino Linotype', 15)
            
        canvas.line(31.83,133,765.92,133) #"""bottom most line"""
        canvas.line(31.83,133,31.83,419) #"""left most line"""
        canvas.line(31.83,419,765.92,419) #"""top most line"""
        canvas.line(765.92,133,765.92,419) #"""right most line"""
        canvas.line(80.23,133,80.23,419) #"""1st column line"""
        canvas.line(179.59,133,179.59,419) #"""2nd column line"""
        canvas.line(226.39,133,226.39,419) #"""3rd column line"""
        canvas.line(324.29,133,324.29,419) #""" 4th column line"""
        canvas.line(371.09,133,371.09,419) #""" 5th column line"""
        canvas.line(473.29,133,473.29,419) #""" 6th column line"""
        canvas.line(520.09,133,520.09,419) #""" 7th column line"""
        canvas.line(621.67,133,621.67,419) #""" 8th column line"""
        canvas.line(667.03,133,667.03,419) #""" 9th column line"""
            
        canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
        canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
        canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
        canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
        canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
            
        canvas.drawString(36.83,401.67,'Table')
        canvas.drawString(42.83,386.67,'No.')
        canvas.drawString(103.64,394.17,'Reg.No')
        canvas.drawString(184.59,401.67,'Table')
        canvas.drawString(190.59,386.67,'No.')
        canvas.drawString(249.8,394.17,'Reg.No')
        canvas.drawString(329.29,401.67,'Table')
        canvas.drawString(335.29,386.67,'No.')
        canvas.drawString(394.5,394.17,'Reg.No')
        canvas.drawString(478.29,401.67,'Table')
        canvas.drawString(484.29,386.67,'No.')
        canvas.drawString(543.5,394.67,'Reg.No')
        canvas.drawString(626.67,401.67,'Table')
        canvas.drawString(632.67,386.67,'No.')
        canvas.drawString(690.44,394.67,'Reg.No')
        r1 = [0,1,2,5]
        canvas.drawString(51.83,349.936,'1')
        canvas.drawString(51.83,300.616,'2')
        canvas.drawString(51.83,251.296,'3')
        canvas.drawString(51.83,201.686,'4')
        canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
            
        canvas.drawString(194.59,349.936,'10')
        canvas.drawString(199.59,300.616,'9')
        canvas.drawString(199.59,251.296,'8')
        canvas.drawString(199.59,201.686,'7')
        canvas.drawString(199.59,152.076,'6')
            
        canvas.drawString(339.29,349.936,'11')
        canvas.drawString(339.29,300.616,'12')
        canvas.drawString(339.29,251.296,'13')
        canvas.drawString(339.29,201.686,'14')
        canvas.drawString(339.29,152.076,'15')
            
        canvas.drawString(488.29,349.936,'20')
        canvas.drawString(488.29,300.616,'19')
        canvas.drawString(488.29,251.296,'18')
        canvas.drawString(488.29,201.686,'17')
        canvas.drawString(488.29,152.076,'16')
            
        canvas.drawString(636.67,349.936,'21')
        canvas.drawString(636.67,300.616,'22')
        canvas.drawString(636.67,251.296,'23')
        canvas.drawString(636.67,201.686,'24')
        canvas.drawString(636.67,152.076,'25')
        canvas.drawString(84.846,349.936,str(dept1[0]).decode('utf-8'))
        canvas.drawString(84.846,300.616,str(dept1[1]).decode('utf-8'))
        canvas.drawString(84.846,251.296,str(dept1[2]).decode('utf-8'))
        canvas.drawString(84.846,201.686,str(dept1[3]).decode('utf-8'))
        canvas.drawString(84.846,152.076,str(dept1[4]).decode('utf-8'))
                
        canvas.drawString(231.006,349.936,str(dept1[9]).decode('utf-8'))
        canvas.drawString(231.006,300.616,str(dept1[8]).decode('utf-8'))
        canvas.drawString(231.006,251.296,str(dept1[7]).decode('utf-8'))
        canvas.drawString(231.006,201.686,str(dept1[6]).decode('utf-8'))
        canvas.drawString(231.006,152.076,str(dept1[5]).decode('utf-8'))
                
        canvas.drawString(375.706,349.936,str(dept1[10]).decode('utf-8'))
        canvas.drawString(375.706,300.616,str(dept1[11]).decode('utf-8'))
        canvas.drawString(375.706,251.296,str(dept1[12]).decode('utf-8'))
        canvas.drawString(375.706,201.686,str(dept1[13]).decode('utf-8'))
        canvas.drawString(375.706,152.076,str(dept1[14]).decode('utf-8'))
                
        canvas.drawString(524.706,349.936,str(dept1[19]).decode('utf-8'))
        canvas.drawString(524.706,300.616,str(dept1[18]).decode('utf-8'))
        canvas.drawString(524.706,251.296,str(dept1[17]).decode('utf-8'))
        canvas.drawString(524.706,201.686,str(dept1[16]).decode('utf-8'))
        canvas.drawString(524.706,152.076,str(dept1[15]).decode('utf-8'))
                
        canvas.drawString(671.646,349.936,str(dept1[20]).decode('utf-8'))
        canvas.drawString(671.646,300.616,str(dept1[21]).decode('utf-8'))
        canvas.drawString(671.646,251.296,str(dept1[22]).decode('utf-8'))
        canvas.drawString(671.646,201.686,str(dept1[23]).decode('utf-8'))
        canvas.drawString(671.646,152.076,str(dept1[24]).decode('utf-8'))
        canvas.showPage()
        hall_index += 1
        hall_no = hall_list[hall_index]
        hall_n()
        hallno = "Hall No - %s %s" % (hall_no,floor)
        ##print "i-%d" % (i)
        i = i + 1
        try:
            
            for s in range(25):
                dept1.pop(0)
        except IndexError:
            pass
        if i == l:
            k = 0
            for k in range(25-len(dept1)):
                dept1.append("")
            pdfmetrics.registerFont(TTFont('Bookman Old Style', 'BOOKOSB.ttf'))
            pdfmetrics.registerFont(TTFont('Palatino Linotype', 'palatino-linotype-bold-italic.ttf'))
            pdfmetrics.registerFont(TTFont('Times New Roman', 'Times_Normal.ttf'))
            canvas.setFont('Times New Roman', 18)
            canvas.drawString(200.83,575.8,clge_name)
            canvas.drawString(200.83,550.8,"Anna University Examinations - ")
            canvas.drawString(457.83,550.8,month)
                 
            canvas.setFont('Bookman Old Style', 23)
            canvas.drawString(369.67,458.8,hallno)
            canvas.setFont('Palatino Linotype', 18)
            canvas.drawString(456.33,490.40,"Session: ")
            canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                    
            canvas.drawString(90.83,490.40,'Date: ')
            canvas.drawString(150.83,490.40,date)
            canvas.drawString(90.83,458.8,'No. of Candidates: ')
            no_ofstudents = 0
            no_ofstudents = max_ele - (l*25)
            canvas.drawString(250.83,458.8,str(no_ofstudents))
            canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
            canvas.setFont('Palatino Linotype', 15)
                
            canvas.line(31.83,133,765.92,133) #"""bottom most line"""
            canvas.line(31.83,133,31.83,419) #"""left most line"""
            canvas.line(31.83,419,765.92,419) #"""top most line"""
            canvas.line(765.92,133,765.92,419) #"""right most line"""
            canvas.line(80.23,133,80.23,419) #"""1st column line"""
            canvas.line(179.59,133,179.59,419) #"""2nd column line"""
            canvas.line(226.39,133,226.39,419) #"""3rd column line"""
            canvas.line(324.29,133,324.29,419) #""" 4th column line"""
            canvas.line(371.09,133,371.09,419) #""" 5th column line"""
            canvas.line(473.29,133,473.29,419) #""" 6th column line"""
            canvas.line(520.09,133,520.09,419) #""" 7th column line"""
            canvas.line(621.67,133,621.67,419) #""" 8th column line"""
            canvas.line(667.03,133,667.03,419) #""" 9th column line"""
                
            canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
            canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
            canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
            canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
            canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
                
            canvas.drawString(36.83,401.67,'Table')
            canvas.drawString(42.83,386.67,'No.')
            canvas.drawString(103.64,394.17,'Reg.No')
            canvas.drawString(184.59,401.67,'Table')
            canvas.drawString(190.59,386.67,'No.')
            canvas.drawString(249.8,394.17,'Reg.No')
            canvas.drawString(329.29,401.67,'Table')
            canvas.drawString(335.29,386.67,'No.')
            canvas.drawString(394.5,394.17,'Reg.No')
            canvas.drawString(478.29,401.67,'Table')
            canvas.drawString(484.29,386.67,'No.')
            canvas.drawString(543.5,394.67,'Reg.No')
            canvas.drawString(626.67,401.67,'Table')
            canvas.drawString(632.67,386.67,'No.')
            canvas.drawString(690.44,394.67,'Reg.No')
            r1 = [0,1,2,5]
            canvas.drawString(51.83,349.936,'1')
            canvas.drawString(51.83,300.616,'2')
            canvas.drawString(51.83,251.296,'3')
            canvas.drawString(51.83,201.686,'4')
            canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
                
            canvas.drawString(194.59,349.936,'10')
            canvas.drawString(199.59,300.616,'9')
            canvas.drawString(199.59,251.296,'8')
            canvas.drawString(199.59,201.686,'7')
            canvas.drawString(199.59,152.076,'6')
                
            canvas.drawString(339.29,349.936,'11')
            canvas.drawString(339.29,300.616,'12')
            canvas.drawString(339.29,251.296,'13')
            canvas.drawString(339.29,201.686,'14')
            canvas.drawString(339.29,152.076,'15')
                
            canvas.drawString(488.29,349.936,'20')
            canvas.drawString(488.29,300.616,'19')
            canvas.drawString(488.29,251.296,'18')
            canvas.drawString(488.29,201.686,'17')
            canvas.drawString(488.29,152.076,'16')
                
            canvas.drawString(636.67,349.936,'21')
            canvas.drawString(636.67,300.616,'22')
            canvas.drawString(636.67,251.296,'23')
            canvas.drawString(636.67,201.686,'24')
            canvas.drawString(636.67,152.076,'25')
            canvas.drawString(84.846,349.936,str(dept1[0]).decode('utf-8'))
            canvas.drawString(84.846,300.616,str(dept1[1]).decode('utf-8'))
            canvas.drawString(84.846,251.296,str(dept1[2]).decode('utf-8'))
            canvas.drawString(84.846,201.686,str(dept1[3]).decode('utf-8'))
            canvas.drawString(84.846,152.076,str(dept1[4]).decode('utf-8'))
                    
            canvas.drawString(231.006,349.936,str(dept1[9]).decode('utf-8'))
            canvas.drawString(231.006,300.616,str(dept1[8]).decode('utf-8'))
            canvas.drawString(231.006,251.296,str(dept1[7]).decode('utf-8'))
            canvas.drawString(231.006,201.686,str(dept1[6]).decode('utf-8'))
            canvas.drawString(231.006,152.076,str(dept1[5]).decode('utf-8'))
                    
            canvas.drawString(375.706,349.936,str(dept1[10]).decode('utf-8'))
            canvas.drawString(375.706,300.616,str(dept1[11]).decode('utf-8'))
            canvas.drawString(375.706,251.296,str(dept1[12]).decode('utf-8'))
            canvas.drawString(375.706,201.686,str(dept1[13]).decode('utf-8'))
            canvas.drawString(375.706,152.076,str(dept1[14]).decode('utf-8'))
                    
            canvas.drawString(524.706,349.936,str(dept1[19]).decode('utf-8'))
            canvas.drawString(524.706,300.616,str(dept1[18]).decode('utf-8'))
            canvas.drawString(524.706,251.296,str(dept1[17]).decode('utf-8'))
            canvas.drawString(524.706,201.686,str(dept1[16]).decode('utf-8'))
            canvas.drawString(524.706,152.076,str(dept1[15]).decode('utf-8'))
                    
            canvas.drawString(671.646,349.936,str(dept1[20]).decode('utf-8'))
            canvas.drawString(671.646,300.616,str(dept1[21]).decode('utf-8'))
            canvas.drawString(671.646,251.296,str(dept1[22]).decode('utf-8'))
            canvas.drawString(671.646,201.686,str(dept1[23]).decode('utf-8'))
            canvas.drawString(671.646,152.076,str(dept1[24]).decode('utf-8'))
            canvas.showPage()
            
                
        
    #canvas.save()


#single_dept(canvas)
#canvas = canvas.Canvas("wtf.pdf", pagesize=landscape(letter))
def multidept(hall_no,canvas):
    #canvas = canvas.Canvas("wtf.pdf", pagesize=landscape(letter))
    with open('college_name.txt', 'r') as myfile:
        clge_name=myfile.read().replace('\n', '')
    clge_name = str(clge_name)
    dept1 = []
    book1 = load_workbook("output1.xlsx")
    dept_01 = book1.get_active_sheet()
    rowct1 = 6
    i1 = 6
    loop1 = 0
    sheet1 = book1.worksheets[0]
    row_count1 = sheet1.max_row
    hall_index = 0
    hall_index = hall_list.index(hall_no)
    hall_n()
    hallno = "Hall No - %s %s" % (hall_no,floor)
    while rowct1 < row_count1:
        try:
            while str(dept_01.cell(row = i1, column = 1).value[0:7]) != "College":
                i1 += 1
                loop1 += 1
                
        except TypeError:
            pass
            
        #print loop1
        r1 = 0
        for r1 in range(loop1):
            if len(str(dept_01.cell(row = rowct1, column = 1).value)) == 14:
                dept1.append(str(dept_01.cell(row = rowct1, column = 1).value[2:14]))
            elif len(str(dept_01.cell(row = rowct1, column = 1).value)) == 15:
                dept1.append(str(dept_01.cell(row = rowct1, column = 1).value[3:15]))
            else:
                dept1.append(str(dept_01.cell(row = rowct1, column = 1).value[4:16]))
            rowct1 = rowct1 + 1
        
        rowct1 = rowct1 + 5
        i1 = rowct1
        loop1 = 0
    ########
    dept2 = []
    book2 = load_workbook("output2.xlsx")
    dept_02 = book2.get_active_sheet()
    rowct2 = 6
    i2 = 6
    loop2 = 0
    sheet2 = book2.worksheets[0]
    row_count2 = sheet2.max_row
    #month = "NOV/DEC"
    #date = "10-12-17"
    #hallno = "IN2512"
    #session = "FN"
    while rowct2 < row_count2:
        try:
            while str(dept_02.cell(row = i2, column = 1).value[0:7]) != "College":
                i2 += 1
                loop2 += 1
                
        except TypeError:
            pass
            
        #print loop2
        r1 = 0
        for r1 in range(loop2):
            if len(str(dept_02.cell(row = rowct2, column = 1).value)) == 14:
                dept2.append(str(dept_02.cell(row = rowct2, column = 1).value[2:14]))
            elif len(str(dept_02.cell(row = rowct2, column = 1).value)) == 15:
                dept2.append(str(dept_02.cell(row = rowct2, column = 1).value[3:15]))
            else:
                dept2.append(str(dept_02.cell(row = rowct2, column = 1).value[4:16]))
            rowct2 = rowct2 + 1
        
        rowct2 = rowct2 + 5
        i2 = rowct2
        loop2 = 0
    #print "len of dept1-%d" % (len(dept1))
    #print "len of dept2-%d" % (len(dept2))
    l1 = len(dept1)/25
    l2 = len(dept2)/25
    try:
        if l1 == l2:
            i = 0
            while i != l1:
                pdfmetrics.registerFont(TTFont('Bookman Old Style', 'BOOKOSB.ttf'))
                pdfmetrics.registerFont(TTFont('Palatino Linotype', 'palatino-linotype-bold-italic.ttf'))
                pdfmetrics.registerFont(TTFont('Times New Roman', 'Times_Normal.ttf'))
                canvas.setFont('Times New Roman', 18)
                canvas.drawString(200.83,575.8,clge_name)
                canvas.drawString(200.83,550.8,"Anna University Examinations - ")
                canvas.drawString(457.83,550.8,month)

                canvas.setFont('Bookman Old Style', 23)
                canvas.drawString(369.67,458.8,hallno)
                canvas.setFont('Palatino Linotype', 18)
                canvas.drawString(456.33,490.40,"Session: ")
                canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                
                canvas.drawString(90.83,490.40,'Date: ')
                canvas.drawString(150.83,490.40,date)
                canvas.drawString(90.83,458.8,'No. of Candidates: ')
                canvas.drawString(250.83,458.8,"25")
                canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
                canvas.setFont('Palatino Linotype', 15)

                
                canvas.line(31.83,133,765.92,133) #"""bottom most line"""
                canvas.line(31.83,133,31.83,419) #"""left most line"""
                canvas.line(31.83,419,765.92,419) #"""top most line"""
                canvas.line(765.92,133,765.92,419) #"""right most line"""
                canvas.line(80.23,133,80.23,419) #"""1st column line"""
                canvas.line(179.59,133,179.59,419) #"""2nd column line"""
                canvas.line(226.39,133,226.39,419) #"""3rd column line"""
                canvas.line(324.29,133,324.29,419) #""" 4th column line"""
                canvas.line(371.09,133,371.09,419) #""" 5th column line"""
                canvas.line(473.29,133,473.29,419) #""" 6th column line"""
                canvas.line(520.09,133,520.09,419) #""" 7th column line"""
                canvas.line(621.67,133,621.67,419) #""" 8th column line"""
                canvas.line(667.03,133,667.03,419) #""" 9th column line"""
                
                canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
                canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
                canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
                canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
                canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
                
                canvas.drawString(36.83,401.67,'Table')
                canvas.drawString(42.83,386.67,'No.')
                canvas.drawString(103.64,394.17,'Reg.No')
                canvas.drawString(184.59,401.67,'Table')
                canvas.drawString(190.59,386.67,'No.')
                canvas.drawString(249.8,394.17,'Reg.No')
                canvas.drawString(329.29,401.67,'Table')
                canvas.drawString(335.29,386.67,'No.')
                canvas.drawString(394.5,394.17,'Reg.No')
                canvas.drawString(478.29,401.67,'Table')
                canvas.drawString(484.29,386.67,'No.')
                canvas.drawString(543.5,394.67,'Reg.No')
                canvas.drawString(626.67,401.67,'Table')
                canvas.drawString(632.67,386.67,'No.')
                canvas.drawString(690.44,394.67,'Reg.No')
                r1 = [0,1,2,5]
                canvas.drawString(51.83,349.936,'1')
                canvas.drawString(51.83,300.616,'2')
                canvas.drawString(51.83,251.296,'3')
                canvas.drawString(51.83,201.686,'4')
                canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
                
                canvas.drawString(194.59,349.936,'10')
                canvas.drawString(199.59,300.616,'9')
                canvas.drawString(199.59,251.296,'8')
                canvas.drawString(199.59,201.686,'7')
                canvas.drawString(199.59,152.076,'6')
                
                canvas.drawString(339.29,349.936,'11')
                canvas.drawString(339.29,300.616,'12')
                canvas.drawString(339.29,251.296,'13')
                canvas.drawString(339.29,201.686,'14')
                canvas.drawString(339.29,152.076,'15')
                
                canvas.drawString(488.29,349.936,'20')
                canvas.drawString(488.29,300.616,'19')
                canvas.drawString(488.29,251.296,'18')
                canvas.drawString(488.29,201.686,'17')
                canvas.drawString(488.29,152.076,'16')
                
                canvas.drawString(636.67,349.936,'21')
                canvas.drawString(636.67,300.616,'22')
                canvas.drawString(636.67,251.296,'23')
                canvas.drawString(636.67,201.686,'24')
                canvas.drawString(636.67,152.076,'25')
                
                canvas.drawString(84.846,349.936,str(dept1[0]).decode('utf-8'))
                canvas.drawString(84.846,300.616,str(dept2[0]).decode('utf-8'))
                canvas.drawString(84.846,251.296,str(dept1[1]).decode('utf-8'))
                canvas.drawString(84.846,201.686,str(dept2[1]).decode('utf-8'))
                canvas.drawString(84.846,152.076,str(dept1[2]).decode('utf-8'))
                
                canvas.drawString(231.006,349.936,str(dept2[4]).decode('utf-8'))
                canvas.drawString(231.006,300.616,str(dept1[4]).decode('utf-8'))
                canvas.drawString(231.006,251.296,str(dept2[3]).decode('utf-8'))
                canvas.drawString(231.006,201.686,str(dept1[3]).decode('utf-8'))
                canvas.drawString(231.006,152.076,str(dept2[2]).decode('utf-8'))
                
                canvas.drawString(375.706,349.936,str(dept1[5]).decode('utf-8'))
                canvas.drawString(375.706,300.616,str(dept2[5]).decode('utf-8'))
                canvas.drawString(375.706,251.296,str(dept1[6]).decode('utf-8'))
                canvas.drawString(375.706,201.686,str(dept2[6]).decode('utf-8'))
                canvas.drawString(375.706,152.076,str(dept1[7]).decode('utf-8'))
                
                canvas.drawString(524.706,349.936,str(dept2[9]).decode('utf-8'))
                canvas.drawString(524.706,300.616,str(dept1[9]).decode('utf-8'))
                canvas.drawString(524.706,251.296,str(dept2[8]).decode('utf-8'))
                canvas.drawString(524.706,201.686,str(dept1[8]).decode('utf-8'))
                canvas.drawString(524.706,152.076,str(dept2[7]).decode('utf-8'))
                
                canvas.drawString(671.646,349.936,str(dept1[10]).decode('utf-8'))
                canvas.drawString(671.646,300.616,str(dept2[10]).decode('utf-8'))
                canvas.drawString(671.646,251.296,str(dept1[11]).decode('utf-8'))
                canvas.drawString(671.646,201.686,str(dept2[11]).decode('utf-8'))
                canvas.drawString(671.646,152.076,str(dept1[12]).decode('utf-8'))
                canvas.showPage()
                #################
                hall_index += 1
                hall_no = hall_list[hall_index]
                hall_n()
                hallno = "Hall No - %s %s" % (hall_no,floor)
                canvas.setFont('Times New Roman', 18)
                canvas.drawString(200.83,575.8,clge_name)
                canvas.drawString(200.83,550.8,"Anna University Examinations - ")
                canvas.drawString(457.83,550.8,month)

                canvas.setFont('Bookman Old Style', 23)
                canvas.drawString(369.67,458.8,hallno)
                canvas.setFont('Palatino Linotype', 18)
                canvas.drawString(456.33,490.40,"Session: ")
                canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                
                canvas.drawString(90.83,490.40,'Date: ')
                canvas.drawString(150.83,490.40,date)
                canvas.drawString(90.83,458.8,'No. of Candidates: ')
                canvas.drawString(250.83,458.8,"25")
                canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
                canvas.setFont('Palatino Linotype', 15)

                
                canvas.line(31.83,133,765.92,133) #"""bottom most line"""
                canvas.line(31.83,133,31.83,419) #"""left most line"""
                canvas.line(31.83,419,765.92,419) #"""top most line"""
                canvas.line(765.92,133,765.92,419) #"""right most line"""
                canvas.line(80.23,133,80.23,419) #"""1st column line"""
                canvas.line(179.59,133,179.59,419) #"""2nd column line"""
                canvas.line(226.39,133,226.39,419) #"""3rd column line"""
                canvas.line(324.29,133,324.29,419) #""" 4th column line"""
                canvas.line(371.09,133,371.09,419) #""" 5th column line"""
                canvas.line(473.29,133,473.29,419) #""" 6th column line"""
                canvas.line(520.09,133,520.09,419) #""" 7th column line"""
                canvas.line(621.67,133,621.67,419) #""" 8th column line"""
                canvas.line(667.03,133,667.03,419) #""" 9th column line"""
                
                canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
                canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
                canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
                canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
                canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
                
                canvas.drawString(36.83,401.67,'Table')
                canvas.drawString(42.83,386.67,'No.')
                canvas.drawString(103.64,394.17,'Reg.No')
                canvas.drawString(184.59,401.67,'Table')
                canvas.drawString(190.59,386.67,'No.')
                canvas.drawString(249.8,394.17,'Reg.No')
                canvas.drawString(329.29,401.67,'Table')
                canvas.drawString(335.29,386.67,'No.')
                canvas.drawString(394.5,394.17,'Reg.No')
                canvas.drawString(478.29,401.67,'Table')
                canvas.drawString(484.29,386.67,'No.')
                canvas.drawString(543.5,394.67,'Reg.No')
                canvas.drawString(626.67,401.67,'Table')
                canvas.drawString(632.67,386.67,'No.')
                canvas.drawString(690.44,394.67,'Reg.No')
                r1 = [0,1,2,5]
                canvas.drawString(51.83,349.936,'1')
                canvas.drawString(51.83,300.616,'2')
                canvas.drawString(51.83,251.296,'3')
                canvas.drawString(51.83,201.686,'4')
                canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
                
                canvas.drawString(194.59,349.936,'10')
                canvas.drawString(199.59,300.616,'9')
                canvas.drawString(199.59,251.296,'8')
                canvas.drawString(199.59,201.686,'7')
                canvas.drawString(199.59,152.076,'6')
                
                canvas.drawString(339.29,349.936,'11')
                canvas.drawString(339.29,300.616,'12')
                canvas.drawString(339.29,251.296,'13')
                canvas.drawString(339.29,201.686,'14')
                canvas.drawString(339.29,152.076,'15')
                
                canvas.drawString(488.29,349.936,'20')
                canvas.drawString(488.29,300.616,'19')
                canvas.drawString(488.29,251.296,'18')
                canvas.drawString(488.29,201.686,'17')
                canvas.drawString(488.29,152.076,'16')
                
                canvas.drawString(636.67,349.936,'21')
                canvas.drawString(636.67,300.616,'22')
                canvas.drawString(636.67,251.296,'23')
                canvas.drawString(636.67,201.686,'24')
                canvas.drawString(636.67,152.076,'25')
                
                canvas.drawString(84.846,349.936,str(dept2[12]).decode('utf-8'))
                canvas.drawString(84.846,300.616,str(dept1[13]).decode('utf-8'))
                canvas.drawString(84.846,251.296,str(dept2[13]).decode('utf-8'))
                canvas.drawString(84.846,201.686,str(dept1[14]).decode('utf-8'))
                canvas.drawString(84.846,152.076,str(dept2[14]).decode('utf-8'))
                
                canvas.drawString(231.006,349.936,str(dept1[17]).decode('utf-8'))
                canvas.drawString(231.006,300.616,str(dept2[16]).decode('utf-8'))
                canvas.drawString(231.006,251.296,str(dept1[16]).decode('utf-8'))
                canvas.drawString(231.006,201.686,str(dept2[15]).decode('utf-8'))
                canvas.drawString(231.006,152.076,str(dept1[15]).decode('utf-8'))
                
                canvas.drawString(375.706,349.936,str(dept2[17]).decode('utf-8'))
                canvas.drawString(375.706,300.616,str(dept1[18]).decode('utf-8'))
                canvas.drawString(375.706,251.296,str(dept2[18]).decode('utf-8'))
                canvas.drawString(375.706,201.686,str(dept1[19]).decode('utf-8'))
                canvas.drawString(375.706,152.076,str(dept2[19]).decode('utf-8'))
                
                canvas.drawString(524.706,349.936,str(dept1[22]).decode('utf-8'))
                canvas.drawString(524.706,300.616,str(dept2[21]).decode('utf-8'))
                canvas.drawString(524.706,251.296,str(dept1[21]).decode('utf-8'))
                canvas.drawString(524.706,201.686,str(dept2[20]).decode('utf-8'))
                canvas.drawString(524.706,152.076,str(dept1[20]).decode('utf-8'))
                
                canvas.drawString(671.646,349.936,str(dept2[22]).decode('utf-8'))
                canvas.drawString(671.646,300.616,str(dept1[23]).decode('utf-8'))
                canvas.drawString(671.646,251.296,str(dept2[23]).decode('utf-8'))
                canvas.drawString(671.646,201.686,str(dept1[24]).decode('utf-8'))
                canvas.drawString(671.646,152.076,str(dept2[24]).decode('utf-8'))
                canvas.showPage()
                hall_index += 1
                hall_no = hall_list[hall_index]
                hall_n()
                hallno = "Hall No - %s %s" % (hall_no,floor)
                i = i + 1
                try:
                    
                    for m in range(25):
                        dept1.pop(0)
                    for n in range(25):
                        dept2.pop(0)

                except IndexError:
                    pass
                if i == l1:
                    atlast(dept1,dept2)
        else:
            least = min(l1,l2)
            i = 0
            while i != least:
                pdfmetrics.registerFont(TTFont('Bookman Old Style', 'BOOKOSB.ttf'))
                pdfmetrics.registerFont(TTFont('Palatino Linotype', 'palatino-linotype-bold-italic.ttf'))
                pdfmetrics.registerFont(TTFont('Times New Roman', 'Times_Normal.ttf'))
                canvas.setFont('Times New Roman', 18)
                canvas.drawString(200.83,575.8,clge_name)
                canvas.drawString(200.83,550.8,"Anna University Examinations - ")
                canvas.drawString(457.83,550.8,month)

                canvas.setFont('Bookman Old Style', 23)
                canvas.drawString(369.67,458.8,hallno)
                canvas.setFont('Palatino Linotype', 18)
                canvas.drawString(456.33,490.40,"Session: ")
                canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                
                canvas.drawString(90.83,490.40,'Date: ')
                canvas.drawString(150.83,490.40,date)
                canvas.drawString(90.83,458.8,'No. of Candidates: ')
                canvas.drawString(250.83,458.8,"25")
                canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
                canvas.setFont('Palatino Linotype', 15)

                
                canvas.line(31.83,133,765.92,133) #"""bottom most line"""
                canvas.line(31.83,133,31.83,419) #"""left most line"""
                canvas.line(31.83,419,765.92,419) #"""top most line"""
                canvas.line(765.92,133,765.92,419) #"""right most line"""
                canvas.line(80.23,133,80.23,419) #"""1st column line"""
                canvas.line(179.59,133,179.59,419) #"""2nd column line"""
                canvas.line(226.39,133,226.39,419) #"""3rd column line"""
                canvas.line(324.29,133,324.29,419) #""" 4th column line"""
                canvas.line(371.09,133,371.09,419) #""" 5th column line"""
                canvas.line(473.29,133,473.29,419) #""" 6th column line"""
                canvas.line(520.09,133,520.09,419) #""" 7th column line"""
                canvas.line(621.67,133,621.67,419) #""" 8th column line"""
                canvas.line(667.03,133,667.03,419) #""" 9th column line"""
                
                canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
                canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
                canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
                canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
                canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
                
                canvas.drawString(36.83,401.67,'Table')
                canvas.drawString(42.83,386.67,'No.')
                canvas.drawString(103.64,394.17,'Reg.No')
                canvas.drawString(184.59,401.67,'Table')
                canvas.drawString(190.59,386.67,'No.')
                canvas.drawString(249.8,394.17,'Reg.No')
                canvas.drawString(329.29,401.67,'Table')
                canvas.drawString(335.29,386.67,'No.')
                canvas.drawString(394.5,394.17,'Reg.No')
                canvas.drawString(478.29,401.67,'Table')
                canvas.drawString(484.29,386.67,'No.')
                canvas.drawString(543.5,394.67,'Reg.No')
                canvas.drawString(626.67,401.67,'Table')
                canvas.drawString(632.67,386.67,'No.')
                canvas.drawString(690.44,394.67,'Reg.No')
                r1 = [0,1,2,5]
                canvas.drawString(51.83,349.936,'1')
                canvas.drawString(51.83,300.616,'2')
                canvas.drawString(51.83,251.296,'3')
                canvas.drawString(51.83,201.686,'4')
                canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
                
                canvas.drawString(194.59,349.936,'10')
                canvas.drawString(199.59,300.616,'9')
                canvas.drawString(199.59,251.296,'8')
                canvas.drawString(199.59,201.686,'7')
                canvas.drawString(199.59,152.076,'6')
                
                canvas.drawString(339.29,349.936,'11')
                canvas.drawString(339.29,300.616,'12')
                canvas.drawString(339.29,251.296,'13')
                canvas.drawString(339.29,201.686,'14')
                canvas.drawString(339.29,152.076,'15')
                
                canvas.drawString(488.29,349.936,'20')
                canvas.drawString(488.29,300.616,'19')
                canvas.drawString(488.29,251.296,'18')
                canvas.drawString(488.29,201.686,'17')
                canvas.drawString(488.29,152.076,'16')
                
                canvas.drawString(636.67,349.936,'21')
                canvas.drawString(636.67,300.616,'22')
                canvas.drawString(636.67,251.296,'23')
                canvas.drawString(636.67,201.686,'24')
                canvas.drawString(636.67,152.076,'25')
                
                canvas.drawString(84.846,349.936,str(dept1[0]).decode('utf-8'))
                canvas.drawString(84.846,300.616,str(dept2[0]).decode('utf-8'))
                canvas.drawString(84.846,251.296,str(dept1[1]).decode('utf-8'))
                canvas.drawString(84.846,201.686,str(dept2[1]).decode('utf-8'))
                canvas.drawString(84.846,152.076,str(dept1[2]).decode('utf-8'))
                
                canvas.drawString(231.006,349.936,str(dept2[4]).decode('utf-8'))
                canvas.drawString(231.006,300.616,str(dept1[4]).decode('utf-8'))
                canvas.drawString(231.006,251.296,str(dept2[3]).decode('utf-8'))
                canvas.drawString(231.006,201.686,str(dept1[3]).decode('utf-8'))
                canvas.drawString(231.006,152.076,str(dept2[2]).decode('utf-8'))
                
                canvas.drawString(375.706,349.936,str(dept1[5]).decode('utf-8'))
                canvas.drawString(375.706,300.616,str(dept2[5]).decode('utf-8'))
                canvas.drawString(375.706,251.296,str(dept1[6]).decode('utf-8'))
                canvas.drawString(375.706,201.686,str(dept2[6]).decode('utf-8'))
                canvas.drawString(375.706,152.076,str(dept1[7]).decode('utf-8'))
                
                canvas.drawString(524.706,349.936,str(dept2[9]).decode('utf-8'))
                canvas.drawString(524.706,300.616,str(dept1[9]).decode('utf-8'))
                canvas.drawString(524.706,251.296,str(dept2[8]).decode('utf-8'))
                canvas.drawString(524.706,201.686,str(dept1[8]).decode('utf-8'))
                canvas.drawString(524.706,152.076,str(dept2[7]).decode('utf-8'))
                
                canvas.drawString(671.646,349.936,str(dept1[10]).decode('utf-8'))
                canvas.drawString(671.646,300.616,str(dept2[10]).decode('utf-8'))
                canvas.drawString(671.646,251.296,str(dept1[11]).decode('utf-8'))
                canvas.drawString(671.646,201.686,str(dept2[11]).decode('utf-8'))
                canvas.drawString(671.646,152.076,str(dept1[12]).decode('utf-8'))
                canvas.showPage()
                #################
                hall_index += 1
                hall_no = hall_list[hall_index]
                hall_n()
                hallno = "Hall No - %s %s" % (hall_no,floor)
                canvas.setFont('Times New Roman', 18)
                canvas.drawString(200.83,575.8,clge_name)
                canvas.drawString(200.83,550.8,"Anna University Examinations - ")
                canvas.drawString(457.83,550.8,month)

                canvas.setFont('Bookman Old Style', 23)
                canvas.drawString(369.67,458.8,hallno)
                canvas.setFont('Palatino Linotype', 18)
                canvas.drawString(456.33,490.40,"Session: ")
                canvas.drawString(540.33,490.40,str(session).decode('utf-8'))
                
                canvas.drawString(90.83,490.40,'Date: ')
                canvas.drawString(150.83,490.40,date)
                canvas.drawString(90.83,458.8,'No. of Candidates: ')
                canvas.drawString(250.83,458.8,"25")
                canvas.drawString(505.05,45.97,'Signature of Chief Superintendent')
                canvas.setFont('Palatino Linotype', 15)

                
                canvas.line(31.83,133,765.92,133) #"""bottom most line"""
                canvas.line(31.83,133,31.83,419) #"""left most line"""
                canvas.line(31.83,419,765.92,419) #"""top most line"""
                canvas.line(765.92,133,765.92,419) #"""right most line"""
                canvas.line(80.23,133,80.23,419) #"""1st column line"""
                canvas.line(179.59,133,179.59,419) #"""2nd column line"""
                canvas.line(226.39,133,226.39,419) #"""3rd column line"""
                canvas.line(324.29,133,324.29,419) #""" 4th column line"""
                canvas.line(371.09,133,371.09,419) #""" 5th column line"""
                canvas.line(473.29,133,473.29,419) #""" 6th column line"""
                canvas.line(520.09,133,520.09,419) #""" 7th column line"""
                canvas.line(621.67,133,621.67,419) #""" 8th column line"""
                canvas.line(667.03,133,667.03,419) #""" 9th column line"""
                
                canvas.line(31.83,378.75,765.92,378.75)#"""1st row line"""
                canvas.line(31.83,329.43,765.92,329.43)#"""2nd row line"""
                canvas.line(31.83,280.11,765.92,280.11)#"""3rd row line"""
                canvas.line(31.83,230.50,765.92,230.50)#"""4th row line"""
                canvas.line(31.83,180.89,765.92,180.89)#"""5th row line"""
                
                canvas.drawString(36.83,401.67,'Table')
                canvas.drawString(42.83,386.67,'No.')
                canvas.drawString(103.64,394.17,'Reg.No')
                canvas.drawString(184.59,401.67,'Table')
                canvas.drawString(190.59,386.67,'No.')
                canvas.drawString(249.8,394.17,'Reg.No')
                canvas.drawString(329.29,401.67,'Table')
                canvas.drawString(335.29,386.67,'No.')
                canvas.drawString(394.5,394.17,'Reg.No')
                canvas.drawString(478.29,401.67,'Table')
                canvas.drawString(484.29,386.67,'No.')
                canvas.drawString(543.5,394.67,'Reg.No')
                canvas.drawString(626.67,401.67,'Table')
                canvas.drawString(632.67,386.67,'No.')
                canvas.drawString(690.44,394.67,'Reg.No')
                r1 = [0,1,2,5]
                canvas.drawString(51.83,349.936,'1')
                canvas.drawString(51.83,300.616,'2')
                canvas.drawString(51.83,251.296,'3')
                canvas.drawString(51.83,201.686,'4')
                canvas.drawString(51.83,152.076,str(r1[3]).decode('utf-8'))
                
                canvas.drawString(194.59,349.936,'10')
                canvas.drawString(199.59,300.616,'9')
                canvas.drawString(199.59,251.296,'8')
                canvas.drawString(199.59,201.686,'7')
                canvas.drawString(199.59,152.076,'6')
                
                canvas.drawString(339.29,349.936,'11')
                canvas.drawString(339.29,300.616,'12')
                canvas.drawString(339.29,251.296,'13')
                canvas.drawString(339.29,201.686,'14')
                canvas.drawString(339.29,152.076,'15')
                
                canvas.drawString(488.29,349.936,'20')
                canvas.drawString(488.29,300.616,'19')
                canvas.drawString(488.29,251.296,'18')
                canvas.drawString(488.29,201.686,'17')
                canvas.drawString(488.29,152.076,'16')
                
                canvas.drawString(636.67,349.936,'21')
                canvas.drawString(636.67,300.616,'22')
                canvas.drawString(636.67,251.296,'23')
                canvas.drawString(636.67,201.686,'24')
                canvas.drawString(636.67,152.076,'25')
                
                canvas.drawString(84.846,349.936,str(dept2[12]).decode('utf-8'))
                canvas.drawString(84.846,300.616,str(dept1[13]).decode('utf-8'))
                canvas.drawString(84.846,251.296,str(dept2[13]).decode('utf-8'))
                canvas.drawString(84.846,201.686,str(dept1[14]).decode('utf-8'))
                canvas.drawString(84.846,152.076,str(dept2[14]).decode('utf-8'))
                
                canvas.drawString(231.006,349.936,str(dept1[17]).decode('utf-8'))
                canvas.drawString(231.006,300.616,str(dept2[16]).decode('utf-8'))
                canvas.drawString(231.006,251.296,str(dept1[16]).decode('utf-8'))
                canvas.drawString(231.006,201.686,str(dept2[15]).decode('utf-8'))
                canvas.drawString(231.006,152.076,str(dept1[15]).decode('utf-8'))
                
                canvas.drawString(375.706,349.936,str(dept2[17]).decode('utf-8'))
                canvas.drawString(375.706,300.616,str(dept1[18]).decode('utf-8'))
                canvas.drawString(375.706,251.296,str(dept2[18]).decode('utf-8'))
                canvas.drawString(375.706,201.686,str(dept1[19]).decode('utf-8'))
                canvas.drawString(375.706,152.076,str(dept2[19]).decode('utf-8'))
                
                canvas.drawString(524.706,349.936,str(dept1[22]).decode('utf-8'))
                canvas.drawString(524.706,300.616,str(dept2[21]).decode('utf-8'))
                canvas.drawString(524.706,251.296,str(dept1[21]).decode('utf-8'))
                canvas.drawString(524.706,201.686,str(dept2[20]).decode('utf-8'))
                canvas.drawString(524.706,152.076,str(dept1[20]).decode('utf-8'))
                
                canvas.drawString(671.646,349.936,str(dept2[22]).decode('utf-8'))
                canvas.drawString(671.646,300.616,str(dept1[23]).decode('utf-8'))
                canvas.drawString(671.646,251.296,str(dept2[23]).decode('utf-8'))
                canvas.drawString(671.646,201.686,str(dept1[24]).decode('utf-8'))
                canvas.drawString(671.646,152.076,str(dept2[24]).decode('utf-8'))
                canvas.showPage()
                hall_index += 1
                hall_no = hall_list[hall_index]
                hall_n()
                hallno = "Hall No - %s %s" % (hall_no,floor)
                i = i + 1
                try:
                    
                    for m in range(25):
                        dept1.pop(0)
                    for n in range(25):
                        dept2.pop(0)

                except IndexError:
                    pass
                if i == least:
                    atlast(dept1,dept2)
    except IndexError:
        pass
    current_ha(hall_no)
                
            
    #canvas.save()
#canvas = canvas.Canvas("wtf.pdf", pagesize=landscape(letter))            
def test():
    import sys
    from datetime import timedelta
    from PyQt4 import QtGui, QtCore

    class Calendar(QtGui.QWidget):
        """
        A QCalendarWidget example
        """

        def __init__(self):
            # create GUI
            QtGui.QMainWindow.__init__(self)
            self.setWindowTitle('Calendar widget')
            # Set the window dimensions
            self.resize(300,100)
            
            # vertical layout for widgets
            self.vbox = QtGui.QVBoxLayout()
            self.setLayout(self.vbox)
            
            # Create a calendar widget and add it to our layout
            self.cal = QtGui.QCalendarWidget()
            self.vbox.addWidget(self.cal)

            # Create a label which we will use to show the date a week from now
            self.lbl = QtGui.QLabel()
            self.vbox.addWidget(self.lbl)
            
            # Connect the clicked signal to the centre handler
            self.connect(self.cal, QtCore.SIGNAL('selectionChanged()'), self.date_changed)

        def date_changed(self):
            """
            Handler called when the date selection has changed
            """
            global date
            # Fetch the currently selected date, this is a QDate object
            date = self.cal.selectedDate()
            # This is a gives us the date contained in the QDate as a native
            # python date[time] object
            pydate = date.toPyDate()
            # Calculate the date a week from now
            sevendays = timedelta(days=7)
            aweeklater = pydate + sevendays
            # Show this date in our label
            self.lbl.setText('Selected Date is: %s' % pydate)
            date = str(pydate)
            l = []
            l = date 
            date = l[8] + l[9] + l[7] + l[5] + l[6] + l[4] + l[0] + l[1] + l[2] + l[3]
            #print date
            date = str(date)
            
    


    # If the program is run directly or passed as an argument to the python
    # interpreter then create a Calendar instance and show it
    if __name__ == "__main__":
        app = QtGui.QApplication(sys.argv)
        gui = Calendar()
        gui.show()
        app.exec_()
def CurSelet1(event):
    widget1 = event.widget
    selection1=widget1.curselection()
    global session
    session = widget.get(selection[0])
    session = str(session)
def CurSelet2(event):
    widget2 = event.widget
    selection2=widget2.curselection()
    global hall_no
    hall_no = widget.get(selection2[0])
    hall_no = str(session)


    
put = "yes"
r = 0
hi = 0
while put == "yes":
    r = r + 1
    def pop():
        global app
        app = Tk()
        app.title("POPUP")
        app.geometry("340x150")
        label = Label(app, text="DO YOU WANT TO CONTINUE", height=0, width=50).place(x=.6,y=20)
        button1 = Button(app, text="YES", width=10, command=b1).place(x=70,y=80)
        button2 = Button(app, text="NO", width=10, command=b2).place(x=180,y=80)
        app.mainloop()

    def dept1():
        global path1
        path1 = askopenfilename()
        convert_csv_to_xlsx1(path1)
        
    def dept2():
        global path2
        path2 = askopenfilename()
        if path2 == path1:
            path2 = ""
        else:
            convert_csv_to_xlsx2(path2)
    def b1():
        global put
        put = "yes"
        app.destroy()
        mgui.destroy()
    def b2():
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, landscape
        canvas = canvas.Canvas("%s-%s(rest)-%d.pdf" % (date,session,r), pagesize=landscape(letter))
        if path2 == "":
            pass
        else:
            atlastmerge(canvas)
            canvas.save()
        app.destroy()
        mgui.destroy()
        
        global put
        if put == "yes":
            put = ""
        else:
            put = ""
    def convert_csv_to_xlsx1(path1):
        import os
        import csv
        import sys

        from openpyxl import Workbook

        #reload(sys)
        #sys.setdefaultencoding('utf8')
        workbook = Workbook()
        worksheet = workbook.active
        with open(path1, 'r') as f:
            reader = csv.reader(f)
            for r, row in enumerate(reader):
                for c, col in enumerate(row):
                    for idx, val in enumerate(col.split(',')):
                        cell = worksheet.cell(row=r+1, column=c+1)
                        cell.value = val
        workbook.save('output1.xlsx')


    
    def convert_csv_to_xlsx2(path2):
        import os
        import csv
        import sys

        from openpyxl import Workbook

        #reload(sys)
        #sys.setdefaultencoding('utf8')
        workbook = Workbook()
        worksheet = workbook.active
        with open(path2, 'r') as f:
            reader = csv.reader(f)
            for r, row in enumerate(reader):
                for c, col in enumerate(row):
                    for idx, val in enumerate(col.split(',')):
                        cell = worksheet.cell(row=r+1, column=c+1)
                        cell.value = val
        workbook.save('output2.xlsx')

       
        
        
        

    def merge():
        #print "called"
        #global date
        global month
        global session
        global hall_no
        global hallno
        global hall_index
        month = str2.get()
        #print month
        session = str3.get()
        ##print session
        hall_no = str4.get()
        ##print hall_no
        #date = str(date)
        month = str(month)
        session = str(session)
        hall_no = str(hall_no)
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter, landscape
        canvas = canvas.Canvas("%s(%s)-%d.pdf" % (date,session,r), pagesize=landscape(letter))
        if path2 == "":
            #print "single"
            single_dept(hall_no,canvas)
            canvas.save()
        else:
            #print "function"
            multidept(hall_no,canvas)
            canvas.save()

        pop()

        


    ##print "done"
    global mgui
    mgui = Tk()
    str1 = StringVar()
    str2 = StringVar()
    str3 = StringVar()
    str4 = StringVar()
    #from PIL import Image
    from Tkinter import Tk, Frame, BOTH, Canvas
    
    #from PIL import Image, ImageTk
    from Tkinter import *
    import PIL.Image,PIL.ImageTk
    import tkFont
    
    import ttk
    style = ttk.Style()
    mgui.geometry( "600x400" )
    mgui.title("Exam Scheduler")
    mgui.state('zoomed')
    #im = PIL.Image.open('todo1.jpg')
    #im = im.resize((1400, 700))
    #photo = PIL.ImageTk.PhotoImage(im)
    #myvar = Label(mgui,image = photo)
    #myvar.place(x=0, y=0, relwidth=1, relheight=1)
    #mlabel = Label(mgui,text="DATE",fg="black").place(x=40,y=75)
    mlabel = Label(mgui,text="MONTH",fg="black").place(x=346,y=468)
    mlabel = Label(mgui,text="SESSION",fg="black").place(x=950,y=468)
    #mlabel = Label(mgui,text="LOCATION",fg="black").place(x=40,y=245)
    mlabel = Label(mgui,text="HALL_NO",fg="black").place(x=950,y=143)
    mbutton1 = Button(mgui,text="MERGE",fg="#f87305",width="15",height="3",command = merge).place(x=620,y=314)
    mbutton2 = Button(mgui,text="DEPT1",fg="#f87305",width="10",height="1",command = dept1).place(x=338,y=321)
    mbutton3 = Button(mgui,text="DEPT2",fg="#f87305",width="10",height="1",command = dept2).place(x=946,y=320)
    mutton4 = Button(mgui,text="Choose",fg="#f87305",width="7",height="1",command=test).place(x=357,y=165)
    mtext2 = Entry(mgui,width="14",textvariable=str2).place(x=346,y=493)
    mtext3 = OptionMenu(mgui,str3,"FN","AN").place(x=950,y=493)
    with open('college_hall_list.txt', 'r') as myfile:
        clge_hall=myfile.read().replace('\n', '')
    clge_hall = str(clge_hall)
    hall_list = clge_hall.split(" ")
    #hall_list.insert(0,"")
    mtext4 = OptionMenu(mgui,str4,*hall_list).place(x=950,y=168)


    mgui.mainloop()
    hi = hi + 1




            
